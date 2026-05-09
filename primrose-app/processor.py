"""
PRIM'ROSE RH — Processeur de fichiers Excel BeeOne
Lit les fichiers paie + production et calcule tous les KPIs
"""
import pandas as pd
import numpy as np
import json
import datetime
from collections import defaultdict
from pathlib import Path

CAMP_START    = {'C2425': (2024, 10), 'C2526': (2025, 10)}
MOIS_FR       = ['Oct','Nov','Dec','Jan','Fev','Mar','Avr','Mai','Jun','Jul','Aou','Sep']
SECTS_OE      = ['SECTEUR 02','SECTEUR 04','SECTEUR 05','SECTEUR 06','SECTEUR 07','SECTEUR 08','SECTEUR 10']
MOIS_MAP      = {
    'oct':10,'nov':11,'dec':12,'janv':1,'fevr':2,'mars':3,'avr':4,
    'mai':5,'juin':6,'juil':7,'aout':8,'sept':9,
    'jan':1,'fev':2,'mar':3,'jui':6,'aou':8,'sep':9,
}

# ── HELPERS ──────────────────────────────────────────────────────────────────
def mois_label(camp: str, qn: int) -> str:
    mi   = (qn - 1) // 2
    half = (qn - 1) % 2
    sy, sm = CAMP_START.get(camp, (2024, 10))
    tm = sy * 12 + sm + mi
    yr = tm // 12
    mo = tm % 12
    if mo == 0: mo = 12; yr -= 1
    return f"{MOIS_FR[mi % 12]} {yr} {'1re' if half == 0 else '2e'}"

def parse_mois_label(lbl: str):
    """Parse 'DD-mois' -> (day, month_num)"""
    s = str(lbl).strip().lower()
    for old, new in [('é','e'),('è','e'),('ê','e'),('û','u'),('ô','o')]:
        s = s.replace(old, new)
    parts = s.split('-')
    if len(parts) != 2:
        return None
    try:
        day = int(parts[0].strip())
        mo_str = parts[1].strip()
        mo = MOIS_MAP.get(mo_str) or MOIS_MAP.get(mo_str[:4]) or MOIS_MAP.get(mo_str[:3])
        return (day, mo) if mo else None
    except:
        return None

def get_qu(camp_yr, camp_mo, year, mo, day):
    ms = (year - camp_yr) * 12 + (mo - camp_mo)
    return ms * 2 + (0 if day <= 15 else 1) + 1

# ── LECTURE PAIE ─────────────────────────────────────────────────────────────
def read_paie(path: str) -> pd.DataFrame:
    paie = pd.read_excel(path, sheet_name='Etat de paie', dtype={'Matricules': str})
    paie['Matricules'] = paie['Matricules'].str.strip().str.zfill(5)
    paie = paie.dropna(subset=['Campagne', 'Quinzaine', 'Mois', 'Année'])
    paie['Campagne']  = paie['Campagne'].str.strip()
    paie['Quinzaine'] = paie['Quinzaine'].str.strip()
    paie['Mois']      = paie['Mois'].astype(int)
    paie['Année']     = paie['Année'].astype(int)
    paie['qu_num']    = paie['Quinzaine'].str[2:].astype(int)
    return paie

# ── LECTURE POINTAGE (spec/sect/op par mat) ───────────────────────────────────
def read_pointage(path: str) -> pd.DataFrame:
    try:
        pt = pd.read_excel(path, sheet_name='Table de Pointage',
                           dtype={'Matricules': str}, usecols=[0, 1, 3, 4, 5, 6, 7])
        pt.columns = ['camp', 'date', 'mat', 'nom', 'op', 'spec', 'sect']
        pt['mat']  = pt['mat'].str.strip().str.zfill(5)
        pt['spec'] = pt['spec'].fillna('').str.strip()
        pt['spec'] = pt['spec'].apply(lambda x: 'Oeillet' if 'illet' in x else x)
        pt['sect'] = pt['sect'].fillna('').str.strip()
        return pt
    except Exception:
        return pd.DataFrame(columns=['camp','date','mat','nom','op','spec','sect'])

# ── LECTURE PRODUCTION ────────────────────────────────────────────────────────
def read_production(path: str) -> dict:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    # Chercher l'onglet Production
    sheet_name = next((s for s in wb.sheetnames if 'prod' in s.lower()), wb.sheetnames[0])
    ws = wb[sheet_name]
    all_rows = [[ws.cell(r, c).value for c in range(1, 10)] for r in range(1, ws.max_row + 1)]
    wb.close()

    # Trouver séparateur C2526
    c2526_start = None
    for i, row in enumerate(all_rows):
        v = str(row[0]).strip() if row[0] else ''
        if '25-26' in v or 'CAMP25' in v:
            c2526_start = i
            break

    def parse_camp(start, end, cy, cm):
        by_qu = defaultdict(lambda: {s: 0 for s in SECTS_OE + ['TOT']})
        cur_y = cy
        for r in all_rows[start:end]:
            lbl = r[0]
            if lbl is None:
                continue
            if isinstance(lbl, datetime.datetime):
                d   = lbl.date()
                qn  = get_qu(cy, cm, d.year, d.month, d.day)
                qk  = f'Q{qn}'
                for si, s in enumerate(SECTS_OE, 1):
                    by_qu[qk][s] += int(r[si] or 0)
                by_qu[qk]['TOT'] += int(r[8] or 0)
                continue
            ls = str(lbl).strip()
            if ls in ['2024', '2025', '2026']:
                cur_y = int(ls); continue
            if any(x in ls for x in ['Total', 'Étiquettes', 'CAMP']):
                continue
            parsed = parse_mois_label(ls)
            if not parsed:
                continue
            day, mo = parsed
            qn = get_qu(cy, cm, cur_y, mo, day)
            qk = f'Q{qn}'
            for si, s in enumerate(SECTS_OE, 1):
                by_qu[qk][s] += int(r[si] or 0)
            by_qu[qk]['TOT'] += int(r[8] or 0)
        return {k: dict(v) for k, v in by_qu.items()}

    prod = {
        'C2425': parse_camp(0, c2526_start if c2526_start else len(all_rows), 2024, 10),
        'C2526': parse_camp(c2526_start if c2526_start else len(all_rows), len(all_rows) - 1, 2025, 10)
    }
    return prod

# ── CONSTRUCTION DATA DASHBOARD ───────────────────────────────────────────────
def build_data(paie: pd.DataFrame, mat_info: pd.DataFrame) -> dict:
    mat_spec = dict(zip(mat_info['mat'], mat_info['spec']))
    mat_sect = dict(zip(mat_info['mat'], mat_info['sect']))
    mat_op   = dict(zip(mat_info['mat'], mat_info['op']))

    paie['spec'] = paie['Matricules'].map(mat_spec).fillna('SG')
    paie['sect'] = paie['Matricules'].map(mat_sect).fillna('')

    # Colonnes paie
    col_net    = 'Net à payer '
    col_anc    = 'Prime Anc'
    col_autres = 'Autres primes'
    col_jc     = 'JC'
    col_cnss   = 'Cotis CNSS'
    col_amo    = 'Cotis AMO'
    col_jr     = 'Nbr Jr'

    data = {}
    for camp in ['C2425', 'C2526']:
        data[camp] = {}
        sub = paie[paie['Campagne'] == camp]
        for qu_num in sorted(sub['qu_num'].unique()):
            q  = sub[sub['qu_num'] == qu_num]
            qk = f'Q{int(qu_num)}'
            nb    = q['Matricules'].nunique()
            net   = float(q[col_net].sum())
            jours = float(q[col_jr].sum())
            tp    = round(jours / (nb * 13) * 100, 1) if nb > 0 else 0

            q_entry = {
                'has_paie': True,
                'mois': mois_label(camp, int(qu_num)),
                'nb': int(nb), 'net': round(net), 'jours': round(jours),
                'anc': round(float(q[col_anc].sum())),
                'autres': round(float(q[col_autres].sum())),
                'jc': round(float(q[col_jc].sum())),
                'tot_primes': round(float((q[col_anc]+q[col_autres]+q[col_jc]).sum())),
                'cnss': round(float(q[col_cnss].sum())),
                'amo': round(float(q[col_amo].sum())),
                'tp': tp,
                'specs': {}
            }

            for sp in sorted(q['spec'].unique()):
                sp_q  = q[q['spec'] == sp]
                sp_nb = sp_q['Matricules'].nunique()
                sp_net = float(sp_q[col_net].sum())
                sp_jours = float(sp_q[col_jr].sum())
                sp_tp = round(sp_jours / (sp_nb * 13) * 100, 1) if sp_nb > 0 else 0

                sects_out = {}
                for sect in sorted(sp_q['sect'].dropna().unique()):
                    if not sect: continue
                    s_q  = sp_q[sp_q['sect'] == sect]
                    sects_out[sect] = {
                        'nb': int(s_q['Matricules'].nunique()),
                        'net': round(float(s_q[col_net].sum()))
                    }

                q_entry['specs'][sp] = {
                    'nb': int(sp_nb), 'net': round(sp_net),
                    'anc':    round(float(sp_q[col_anc].sum())),
                    'autres': round(float(sp_q[col_autres].sum())),
                    'jc':     round(float(sp_q[col_jc].sum())),
                    'tot_primes': round(float((sp_q[col_anc]+sp_q[col_autres]+sp_q[col_jc]).sum())),
                    'cnss':   round(float(sp_q[col_cnss].sum())),
                    'amo':    round(float(sp_q[col_amo].sum())),
                    'tp': sp_tp, 'sects': sects_out
                }
            data[camp][qk] = q_entry
    return data

# ── COUT/TIGE ─────────────────────────────────────────────────────────────────
def build_ct(data: dict, prod: dict) -> dict:
    ct = {}
    for camp in ['C2425', 'C2526']:
        ct[camp] = {}
        for qk, q in data[camp].items():
            oe  = q.get('specs', {}).get('Oeillet', {})
            pq  = prod.get(camp, {}).get(qk, {})
            net = oe.get('net', 0)
            tot = pq.get('TOT', 0)
            ct[camp][qk] = round(net / tot, 6) if tot > 0 and net > 0 else 0
    return ct

# ── POINT D'ENTRÉE PRINCIPAL ──────────────────────────────────────────────────
def process_files(paie_path: str, prod_path: str) -> dict:
    """Traite les 2 fichiers Excel et retourne les données du dashboard"""
    print(f"Traitement: {paie_path} + {prod_path}")

    # Lire paie
    paie = read_paie(paie_path)
    print(f"  Paie: {len(paie)} lignes, {paie['Matricules'].nunique()} ouvriers")

    # Lire pointage (spec/sect/op) depuis le même fichier paie
    pt = read_pointage(paie_path)
    if len(pt) > 0:
        mat_spec = pt.groupby('mat')['spec'].agg(lambda x: x.mode()[0] if len(x.mode()) > 0 else 'SG').reset_index()
        mat_sect = pt.groupby('mat')['sect'].agg(lambda x: x.mode()[0] if len(x.mode()) > 0 else '').reset_index()
        mat_op   = pt.groupby('mat')['op'].agg(lambda x: x.mode()[0] if len(x.mode()) > 0 else '').reset_index()
        mat_info = mat_spec.merge(mat_sect, on='mat').merge(mat_op, on='mat')
        mat_info.columns = ['mat', 'spec', 'sect', 'op']
    else:
        # Fallback: assigner SG à tous
        mats = paie['Matricules'].unique()
        mat_info = pd.DataFrame({'mat': mats, 'spec': 'SG', 'sect': '', 'op': ''})

    print(f"  Pointage: {len(pt)} lignes, {mat_info['mat'].nunique()} matricules")

    # Lire production
    prod = read_production(prod_path)
    tot_2425 = sum(v['TOT'] for v in prod.get('C2425', {}).values())
    tot_2526 = sum(v['TOT'] for v in prod.get('C2526', {}).values())
    print(f"  Production: C2425={tot_2425:,} | C2526={tot_2526:,} tiges")

    # Construire data
    data = build_data(paie, mat_info)
    ct   = build_ct(data, prod)

    return {'data': data, 'prod': prod, 'ct': ct}

def load_data(path: str) -> dict:
    """Charge les données sauvegardées"""
    return json.loads(Path(path).read_text())
